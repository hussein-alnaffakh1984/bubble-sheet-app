# -*- coding: utf-8 -*-
"""
مولد التقارير - يصدر النتائج إلى Excel
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ReportGenerator:
    """مولد تقارير Excel"""
    
    def export_excel(self, exam, results, output_path):
        """تصدير النتائج لملف Excel"""
        wb = Workbook()
        
        # ورقة 1: ملخص النتائج
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._create_summary_sheet(ws_summary, exam, results)
        
        # ورقة 2: التفاصيل الكاملة
        ws_details = wb.create_sheet("Details")
        self._create_details_sheet(ws_details, exam, results)
        
        # ورقة 3: تحليل الأسئلة
        ws_analysis = wb.create_sheet("Question Analysis")
        self._create_analysis_sheet(ws_analysis, exam, results)
        
        # ورقة 4: نموذج الإجابة
        ws_key = wb.create_sheet("Answer Key")
        self._create_answer_key_sheet(ws_key, exam)
        
        wb.save(output_path)
        return output_path
    
    def _styled_header(self, cell):
        """تنسيق رأس العمود"""
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    def _styled_cell(self, cell, color=None):
        """تنسيق خلية"""
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        if color:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    def _create_summary_sheet(self, ws, exam, results):
        """ورقة الملخص"""
        # عنوان
        ws['A1'] = f"Exam Results: {exam['title']}"
        ws['A1'].font = Font(bold=True, size=16, color="4472C4")
        ws.merge_cells('A1:G1')
        
        ws['A2'] = f"Subject: {exam.get('subject', '')}"
        ws['A2'].font = Font(italic=True, size=10)
        ws.merge_cells('A2:G2')
        
        # رؤوس
        headers = ['#', 'Student ID', 'Score', 'Percentage', 'Correct', 'Wrong', 'Blank', 'Status']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            self._styled_header(cell)
        
        # البيانات
        successful_results = [r for r in results if r.get('success')]
        successful_results.sort(key=lambda x: x.get('percentage', 0), reverse=True)
        
        for idx, r in enumerate(successful_results, 1):
            row = idx + 4
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=r.get('student_id', '-'))
            ws.cell(row=row, column=3, value=r.get('score', 0))
            ws.cell(row=row, column=4, value=f"{r.get('percentage', 0)}%")
            ws.cell(row=row, column=5, value=r.get('correct', 0))
            ws.cell(row=row, column=6, value=r.get('wrong', 0))
            ws.cell(row=row, column=7, value=r.get('blank', 0))
            
            percentage = r.get('percentage', 0)
            if percentage >= 90:
                status = "Excellent"
                color = "C6EFCE"
            elif percentage >= 75:
                status = "Very Good"
                color = "D9E1F2"
            elif percentage >= 50:
                status = "Pass"
                color = "FFF2CC"
            else:
                status = "Fail"
                color = "FFC7CE"
            
            status_cell = ws.cell(row=row, column=8, value=status)
            for col in range(1, 9):
                self._styled_cell(ws.cell(row=row, column=col), color if col == 8 else None)
        
        # إحصائيات
        if successful_results:
            stats_row = len(successful_results) + 7
            ws.cell(row=stats_row, column=1, value="Statistics").font = Font(bold=True, size=14)
            ws.merge_cells(f'A{stats_row}:H{stats_row}')
            
            scores = [r.get('score', 0) for r in successful_results]
            percentages = [r.get('percentage', 0) for r in successful_results]
            
            stats = [
                ('Total Students', len(successful_results)),
                ('Average Score', round(sum(scores) / len(scores), 2)),
                ('Average Percentage', f"{round(sum(percentages) / len(percentages), 2)}%"),
                ('Maximum Score', max(scores)),
                ('Minimum Score', min(scores)),
                ('Passed (>=50%)', sum(1 for p in percentages if p >= 50)),
                ('Failed (<50%)', sum(1 for p in percentages if p < 50)),
            ]
            
            for i, (label, value) in enumerate(stats):
                ws.cell(row=stats_row + 1 + i, column=1, value=label).font = Font(bold=True)
                ws.cell(row=stats_row + 1 + i, column=2, value=value)
        
        # عرض الأعمدة
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 16
    
    def _create_details_sheet(self, ws, exam, results):
        """ورقة التفاصيل الكاملة"""
        num_q = exam['num_questions']
        
        # رؤوس
        headers = ['#', 'Student ID', 'Score', '%']
        for q in range(1, num_q + 1):
            headers.append(f'Q{q}')
        
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            self._styled_header(cell)
        
        # البيانات
        successful_results = [r for r in results if r.get('success')]
        successful_results.sort(key=lambda x: x.get('student_id', ''))
        
        for idx, r in enumerate(successful_results, 1):
            row = idx + 1
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=r.get('student_id', '-'))
            ws.cell(row=row, column=3, value=r.get('score', 0))
            ws.cell(row=row, column=4, value=f"{r.get('percentage', 0)}%")
            
            detailed = r.get('detailed_results', [])
            for q_idx in range(num_q):
                col = q_idx + 5
                if q_idx < len(detailed):
                    d = detailed[q_idx]
                    if d.get('is_blank'):
                        value = '-'
                        color = "F2F2F2"
                    elif d.get('is_correct'):
                        value = d.get('student_answer', '')
                        color = "C6EFCE"
                    else:
                        value = d.get('student_answer', '')
                        color = "FFC7CE"
                    
                    cell = ws.cell(row=row, column=col, value=value)
                    self._styled_cell(cell, color)
                else:
                    ws.cell(row=row, column=col, value='-')
        
        # عرض الأعمدة
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        for q in range(num_q):
            ws.column_dimensions[get_column_letter(q + 5)].width = 5
    
    def _create_analysis_sheet(self, ws, exam, results):
        """ورقة تحليل الأسئلة"""
        successful_results = [r for r in results if r.get('success')]
        
        # رؤوس
        headers = ['Question', 'Correct Answer', 'Correct Count', 'Wrong Count', 'Blank Count', 'Success Rate', 'Difficulty']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            self._styled_header(cell)
        
        num_q = exam['num_questions']
        answer_key = exam.get('answer_key', [])
        
        for q_idx in range(num_q):
            row = q_idx + 2
            correct = 0
            wrong = 0
            blank = 0
            
            for r in successful_results:
                detailed = r.get('detailed_results', [])
                if q_idx < len(detailed):
                    d = detailed[q_idx]
                    if d.get('is_blank'):
                        blank += 1
                    elif d.get('is_correct'):
                        correct += 1
                    else:
                        wrong += 1
            
            total = len(successful_results)
            success_rate = (correct / total * 100) if total > 0 else 0
            
            if success_rate >= 80:
                difficulty = "Easy"
                color = "C6EFCE"
            elif success_rate >= 50:
                difficulty = "Medium"
                color = "FFF2CC"
            else:
                difficulty = "Hard"
                color = "FFC7CE"
            
            ws.cell(row=row, column=1, value=q_idx + 1)
            ws.cell(row=row, column=2, value=answer_key[q_idx] if q_idx < len(answer_key) else '-')
            ws.cell(row=row, column=3, value=correct)
            ws.cell(row=row, column=4, value=wrong)
            ws.cell(row=row, column=5, value=blank)
            ws.cell(row=row, column=6, value=f"{round(success_rate, 1)}%")
            diff_cell = ws.cell(row=row, column=7, value=difficulty)
            
            for col in range(1, 8):
                self._styled_cell(ws.cell(row=row, column=col), color if col == 7 else None)
        
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 16
    
    def _create_answer_key_sheet(self, ws, exam):
        """ورقة نموذج الإجابة"""
        ws.cell(row=1, column=1, value="Question").font = Font(bold=True)
        ws.cell(row=1, column=2, value="Correct Answer").font = Font(bold=True)
        self._styled_header(ws.cell(row=1, column=1))
        self._styled_header(ws.cell(row=1, column=2))
        
        answer_key = exam.get('answer_key', [])
        for i, ans in enumerate(answer_key):
            ws.cell(row=i + 2, column=1, value=i + 1)
            ws.cell(row=i + 2, column=2, value=ans)
            self._styled_cell(ws.cell(row=i + 2, column=1))
            self._styled_cell(ws.cell(row=i + 2, column=2))
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 18
